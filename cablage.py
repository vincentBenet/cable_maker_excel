import os
import sys
import openpyxl 
import graphviz

os.environ["PATH"] += os.pathsep + r'C:\Program Files\Graphviz\bin'

def extract_from_excel(path_excel, uid="Nom"):
    """
    Extraction des données de excel pour mise dans un dictionnaire avec la clef de l'uid
    """
    print(f"Openning file {path_excel}")
    workbook = openpyxl.load_workbook(path_excel, data_only=True)
    names_sheets = workbook.sheetnames
    print(f"Sheets: {path_excel}")
    data = {}
    data_transformed = {}
    for name_sheet in names_sheets:
        data[name_sheet] = {}
        data_transformed[name_sheet] = {}
        sheet = workbook[name_sheet]
        nb_rows = sheet.max_row
        nb_cols = sheet.max_column
        print(f"Reading sheet {name_sheet}: {nb_rows} rows X {nb_cols} columns")
        names_columns = [header.value for header in sheet[1]]
        for name_column in names_columns:
            data[name_sheet][name_column] = []
        for i_line in range(1, nb_rows):
            values = [line.value for line in sheet[i_line+1]]
            for i_value, value in enumerate(values):
                data[name_sheet][names_columns[i_value]].append(value)
        for i_nom, nom in enumerate(data[name_sheet][uid]):
            if nom in data_transformed[name_sheet]:
                raise Exception(f"Dupplicated entry < {nom} > in < {name_sheet} > at line {i_nom+1}")
            data_transformed[name_sheet][nom] = {}
            for name_column in names_columns:
                data_transformed[name_sheet][nom][name_column] = data[name_sheet][name_column][i_nom]
    print("Finished extracting datas from excel file\n\n")
    return data_transformed


def create_fils(data):
    """
    Foreach connecteur, process the data to know how much connection are going out from connecteurs.
    """
    fils = {}
    for name_connector in data["Connecteurs"]:
        print(f"Connecteur: {name_connector}: ", end="")
        type_connecteur = data["Connecteurs"][name_connector]["_Type connecteur"]
        connection = ""
        interface = ""
        connecteur = ""
        for connection_line in data["Connections"]:
            connecteurs = [
                data["Connections"][connection_line]["From Connecteur"],
                data["Connections"][connection_line]["To Connecteur"]
            ]
            if name_connector in connecteurs:
                data["Connecteurs"][name_connector]["_Cable"] = data["Connections"][connection_line]["_Cable"]
                connection = data["Connections"][connection_line]["Nom"]
                if name_connector == connecteurs[0]:
                    interface = data["Connections"][connection_line]["Interface From"]
                    connecteur = data["Connections"][connection_line]["From Connecteur"]
                if name_connector == connecteurs[1]:
                    interface = data["Connections"][connection_line]["Interface To"]
                    connecteur = data["Connections"][connection_line]["To Connecteur"]
                break
        data["Connecteurs"][connecteur]["Interface"] = interface
        
        fils[connecteur] = {}
        i = 0
        for fil in data["Fils"]:
            if data["Fils"][fil]["_Connecteur"] != connecteur:
                continue
            i += 1
            name_fil = data["Fils"][fil]["_Pin"]
            type_pin = data["Pins"][name_fil]["_Type Pin"]
            fils[connecteur][fil] = {
                "Nom": fil,
                "Label": data["Fils"][fil]["Label"],
                "Couleur": data["Fils"][fil]["Couleur"],
                "Numero": data["Pins"][name_fil]["Numero"],
                "Position": data["Pins"][name_fil]["Label"],
                "PinType": type_pin,
            }
        print(f"{i} fils")
    print("\n")
    return data, fils


def create_graph(path_graph):
    """
    Creating the graphviz graph and its settings
    """
    graph = graphviz.Digraph(
        name='Graph', 
        filename=path_graph, 
        format="svg",
        engine="dot",
        node_attr={
            "shape": "box",
            "peripheries": "1",
            "distortion": "0",
        },
        edge_attr={
            "arrowhead": "none",
            "decorate": "true",
            "labelfloat": "false",
            "labelfontsize": "15",
            "labelfontcolor": "black",
            "portPos": "_",
            "constraint": "true",
            "labeldistance": "5",
        },
        
    )
    graph.attr(**{
        "rankdir": "LR",
        "rank": "same",
        "splines": "ortho",
        "compound": "true",
        "nodesep": "5",
        "ranksep": "5",
        "center": "true",
    })
    
    return graph


def create_cables(graph, data, connecteurs_fils):
    """
    Creating cables subgraphs
    """
    graphs_cables = {}
    for name_cable in data["Cables"]:
        print(f"{name_cable = }")
        
        graphs_cables[name_cable] = graphviz.Digraph(
            name=f"cluster_{name_cable}",
            filename=os.path.join(path_current, "build", f"{data['Cables'][name_cable]['Reference Interne']} - {name_cable.replace('/', '')}"),
            format="svg",
        )
        graphs_cables[name_cable].attr(**{
            "style": 'filled', 
            "concentrate": 'true', 
            "color": 'grey', 
            "label": f"Cable: {name_cable}\nReference Interne: {data['Cables'][name_cable]['Reference Interne']}",
            "rankdir": "LR",
        })
        
    return graphs_cables


def create_connecteurs(data, fils):
    """
    Creating connecteur subgraphs and connecteur node
    """
    graphs_connecteurs = {}
    for name_connecteur in data["Connecteurs"]:
        connecteur = data["Connecteurs"][name_connecteur]
        type_connecteur = connecteur['_Type connecteur']
        print(f"{type_connecteur = }")
        graphs_connecteurs[name_connecteur] = graphviz.Digraph(name=f"cluster_{name_connecteur}")
        graphs_connecteurs[name_connecteur].attr(
            style='filled', 
            splines='ortho', 
            color='lightgrey', 
            label=f"Connecteur: {name_connecteur}"
        )
        graphs_connecteurs[name_connecteur].node(
            name_connecteur,
            label="", 
            image=os.path.join(path_current, "schemas", "Connecteurs", f"{type_connecteur}.svg"),
            xlabel = data["Connecteurs"][name_connecteur]["Interface"],
        )
    return graphs_connecteurs


def add_pins(data, fils, graphs_connecteurs):
    """
    Create pins inside each connecteur subgraphs 
    """
    for name_connector in graphs_connecteurs:
        for name_fil in fils[name_connector]:
            print(f"{name_connector = }")
            print(f"{name_fil = }")
            fil = data["Fils"][name_fil]
            name_pin = fil["_Pin"]
            pin = data["Pins"][name_pin]
            numero = str(pin["Numero"])
            label = str(pin["Label"])
            type_pin = str(pin["_Type Pin"])
            ref_fab = data["Types Pins"][type_pin]["Reference Fabricant"]
            ref_int = data["Types Pins"][type_pin]["Reference Interne"]
            composant = data["Fils"][name_fil]["_Composant"]
            print(f"{composant = }")
            if composant is not None:
                graphs_connecteurs[name_connector].node(
                    f"Composant_{name_fil}", **{
                        "image": os.path.join(path_current, "schemas", "Composants", f"{composant}.svg"),
                        "label": "",
                        "xlabel": f'{composant}: {data["Composants"][composant]["Label"]}'
                    }
                )
            graphs_connecteurs[name_connector].node(name_fil, **{
                "label": f"{numero} : {label}" if numero != label else label,
                "tooltip": f"Pin type: {type_pin}\nLabel: {label}\nNumero: {numero}",
                "shape": "circle",
                "sortv": numero,
                "xlabel": f"{ref_fab if ref_fab is not None else 'MISSING'} / {ref_int if ref_int is not None else 'MISSING'}"
            })


def link_fils(data, fils, graphs_connecteurs):
    """
    Create fils between pins and interface
    """
    # for name_connector in graphs_connecteurs:
        # subgraph = graphs_connecteurs[name_connector]
    for name_connector in fils:
        for name_fil in fils[name_connector]:
            fil = data["Fils"][name_fil]
            name_connector = fil["_Connecteur"]
            subgraph = graphs_connecteurs[name_connector]
            name_pin = fil["_Pin"]
            type_pin = fils[name_connector][name_fil]["PinType"]
            composant = data["Fils"][name_fil]["_Composant"]
            if composant is None:
                subgraph.edge(
                    name_fil, name_connector, **{
                        "color": fil["Couleur"],
                        "xlabel": fil["Label"],
                        "headlabel": fil["Couleur"],
                        "taillabel": type_pin,
                        "penwidth": "2",
                    }
                )
            else:
                subgraph.edge(
                    name_fil, f"Composant_{name_fil}", **{
                        "color": fil["Couleur"],
                        "xlabel": fil["Label"],
                        "headlabel": data["Composants"][composant]["Reference Fabricant"],
                        "taillabel": type_pin,
                        "penwidth": "2",
                    }
                )
                subgraph.edge(
                    f"Composant_{name_fil}", name_connector, **{
                        "color": fil["Couleur"],
                        "xlabel": "",
                        "headlabel": f'fil["Couleur"]',
                        "taillabel": data["Composants"][composant]["Reference Interne"],
                        "penwidth": "2",
                    }
                )


def compose_connecteurs(data, graphs_connecteurs, graphs_cables):
    for name_connector in graphs_connecteurs:
        connecteur = data["Connecteurs"][name_connector]
        name_cable = connecteur["_Cable"]
        subgraph_cable = graphs_cables[name_cable]
        subgraph_conneteur = graphs_connecteurs[name_connector]
        subgraph_cable.subgraph(subgraph_conneteur)


def link_connecteurs(data, graphs_cables):
    for name_connection in data["Connections"]:
        connection = data["Connections"][name_connection]
        name_connecteur_from = connection["From Connecteur"]
        name_connecteur_to = connection["To Connecteur"]
        name_cable = connection["_Cable"]
        type_cable = connection["_Type Cable"]
        ref_interne_cable = data["Types Cables"][type_cable]["Reference Fabricant"]
        ref_externe_cable = data["Types Cables"][type_cable]["Reference Interne"]
        graphs_cables[name_cable].edge(
            name_connecteur_from,
            name_connecteur_to, **{
                "xlabel": f"Label: {connection['Label']}\nLongeur: {connection['Longeur']}",
                "headlabel": "",
                "taillabel": "",
                "penwidth": "8",
                "headport": "c",
                "color": connection['Couleur'],
            }
        )
    return


def compose_cables(data, graph, graphs_cables):
    for name_cable in graphs_cables:
        subgraph_cable = graphs_cables[name_cable]
        graph.subgraph(subgraph_cable)


def make_bom(data):
    pass


def save_graph(graph, graphs_cables):
    graph.render(view=True)
    for name_cable in graphs_cables:
        graph_cable = graphs_cables[name_cable]
        graph_cable.render(view=False)


if __name__ == "__main__":
    # Paths management
    path_current = os.path.dirname(__file__)
    path_excel = os.path.join(path_current, "cablage.xlsx")
    path_graph = os.path.join(path_current, "build", "graph")
    path_test = os.path.join(path_current, "test.png")

    # Data management
    data = extract_from_excel(path_excel)
    data, fils = create_fils(data)
    
    # Graphviz management
    graph = create_graph(path_graph)
    graphs_cables = create_cables(graph, data, fils)
    print(f"{graphs_cables = }")
    graphs_connecteurs = create_connecteurs(data, fils)
    add_pins(data, fils, graphs_connecteurs)
    link_fils(data, fils, graphs_connecteurs)
    compose_connecteurs(data, graphs_connecteurs, graphs_cables)
    link_connecteurs(data, graphs_cables)
    compose_cables(data, graph, graphs_cables)
    make_bom(data)
    save_graph(graph, graphs_cables)
    